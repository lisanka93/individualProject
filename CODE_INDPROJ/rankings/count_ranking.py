# -*- coding: utf-8 -*-

import openpyxl
import os
from collections import Counter



directory = """INSERT HERE THE FOLDER CONTAINING THE SPEADSHEETS WITH DEBATES WITH ARGUMENT PAIRS"""

for file in os.listdir(directory):

	file_name = directory+file
	wb = openpyxl.load_workbook(file_name)
	print file_name
	sheet = wb.get_sheet_by_name('Sheet1')
	end = sheet.max_row
	collection = []

	for i in range(1, end+1, 1):
		if sheet.cell(row = i, column = 3).value == '1,0':
			
			collection.append(sheet.cell(row = i, column = 1).value)
		else:
			
			collection.append(sheet.cell(row = i, column = 2).value)


	wb.save(file_name)


	occurances = Counter(collection)
	print occurances


