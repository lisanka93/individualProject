
#PROCESS DATA
import openpyxl
import os
from whole_debate_func_clean import *
from ind_arguments import*
from collections import Counter

#**********************************************************************************************************************
"""right so this loops through the individual argument folder and analyses the debate and the individual arguments"""
#**********************************************************************************************************************
def demo():
	directory = """FOLDER WITH INDIVIDUAL ARGUMENTS"""
	for file in os.listdir(directory):
		print "******************************************************************************************************************************"
		print "name of file:" , file
		argument_counter = 0


		#reads in excel spreadsheet
		whole_name = directory+file
		wb = openpyxl.load_workbook(whole_name)
		sheet = wb.get_sheet_by_name('Sheet1')
		end = sheet.max_row
		text = []
		


		#let the fun begin: first append all sentences to one big text
		for i in range(2,end,1):
			sentence = [sheet.cell(row = i, column =2).value.encode('utf-8')]
			text = text+sentence
			argument_counter += 1
		
		text= ' '.join(str(r) for r in text)
		#print text                              
		text_length = len(text.split())                   
		#print "text length", text_length
		print "that many arguments", argument_counter
		debate_sentences = len(nltk.sent_tokenize(text))
		#print "that many sentences", debate_sentences
		
		###############################################################################################
		#average length of argument

		average_length = ((float(text_length)) / argument_counter ) 
		print "average length of arguments in words", average_length

		average_sent = ((float(debate_sentences)) / argument_counter)
		print "average length of arguments in sentences", average_sent

		average_sentlength = ((float(text_length)) / debate_sentences)
		print "average sent length", average_sentlength

		#deletes punctuation and hyperlinks, tokenises debate
		tuple_hyp = preprocess_debateI(text)
		
		#number of hyperlinks mentioned in whole debate
		number_hyperlinks = tuple_hyp[1]
		#print "hyperlinks in whole debate: ", number_hyperlinks
		
		average_hyperlink = (float(number_hyperlinks) / argument_counter)
		#print "average number of hyperlinks per argument", average_hyperlink                                        
		no_punct = tuple_hyp[0]
		#print no_punct
		
		#all meaningful/relevant words
		no_stopwords = preprocess_debateII(no_punct)

		#all meaningful words lemmatised
		no_stop_lemmatised = preprocess_debateIII(no_stopwords)

		#long words
		longWords = long_words(no_stop_lemmatised)
		print "long words", longWords

		#"rare" long words with freqdist 1
		very_rare  = very_rare_long_words(no_stop_lemmatised, longWords)
		print "very rare", very_rare
		
		#most common lemmas and stems
		mc_lemma = lemmatised_words(no_stop_lemmatised)
		print "mc lemmas", mc_lemma
		
		mc_stem = stemmed_snowball(no_stopwords)
		print "mc stems", mc_stem
		
		#most common nouns
		nouns = extract_nouns(no_stopwords)
		print "mc nouns", nouns

		bigrams = bigram_extr(no_punct)
		print "mc bigrams", bigrams
		trigrams  = trigram_extr(no_punct)
		print "mc trigrams", trigrams


		#named entities
		named_entities = preprocess_for_nee_and_print(text)
		print "named entities in whole debate", named_entities    
		
		av_ne = round((float(len(named_entities)) / argument_counter),2)
		print "average number of NE mentioned in argument", av_ne              

		#unusual words compared to most common 10000 google words
		print "UNUSUAL WORDS ACCORDING TO MOST COMMONLY USED GOOGLE"
		
		unusual_wordsI = unusual_words(text, mc_stem, google10_stemed)
		print "10000 google words", unusual_wordsI
		
		#unusual words compared to most common 20000 google words
		unusual_wordsII = unusual_words(text, mc_stem, google20_stemed)
		print "20000 google words" ,unusual_wordsII


		#######################################################################################################################
		
		print "_____________________________ individual arguments ____________________________________________________________"
		print "results sentence"
		#prints all the interesting values of each individual argument
		for i in range(2, end+1, 1):

			#sentence aka argument in one cell
			sentence = sheet.cell(row = i, column =2).value.encode('utf-8')
			argu_length = len(sentence)
			print "argulength", argu_length

			#################################################################################
			
			#super preprocessed for mc stuff extraction
			token_sentence = preprocess_argumentIV(sentence)         
			
			#string without hyperlinks
			no_hyperlinks= preprocess_argumentI(sentence)
			
			#delete all punctuatuin apart from apostrophe (also string) 
			only_apostrophe_alllower = preprocess_argumentII(no_hyperlinks)

			#list
			all_lower_no_punct = preprocess_argumentIII(only_apostrophe_alllower)

			#lemmatised sentence
			lemma_sent = [lemma(w).encode('utf-8') for w in token_sentence]
			#print "lemmatised sentence", lemma_sent
			
			#stemmed sentence
			snowball = SnowballStemmer("english")
			
			stem_sent  = [snowball.stem(w).encode('utf-8') for w in token_sentence]
			#print "stemmed stnence", stem_sent
			##################################################################################
			

			nr_sentences = len(nltk.sent_tokenize(sentence))
			print "number of sentences in argument", nr_sentences

			words_argument = len(sentence.split())
			print "words in argument", words_argument



			print "..................................average stuff................................................"

			av_word_sen = round((float(words_argument)/nr_sentences),2)
			print "averagde number of words per sentence in argument", av_word_sen   #0

			av_len_word = round((float(argu_length) / words_argument),2)
			print "average length of word", av_len_word              #1

			#argument longer than average (words)?
			word_ratio = round(float(words_argument) / average_length,2)           #2
			print "word ratio", word_ratio

			#argument longer than average (sentence)?
			sent_ratio = round(float(nr_sentences) / average_sent,2)              #3
			print "sent rati0", sent_ratio

			#sentence length longer than average?
			ratio_ws = round(float(av_word_sen) / average_sentlength,2)       #4
			print "ratio ws", ratio_ws

			print ".................................style........................................................"
			linkwords = linkingwordcount(all_lower_no_punct)                
			linkwords = round((float(linkwords) / nr_sentences),2)
			print " linking words relative to sentence", linkwords         #5
			
			readability = round(ColemanLiauIndex(no_hyperlinks),2)       #6
			print "readability", readability
			
			print "..........................................aggressive/error..................................."


			badword = round(badwordcount(all_lower_no_punct),2)                       
			print "badwords",  badword              #7
			
			spellcheck = round(spellCheck(only_apostrophe_alllower.split()),2)
			print "errors relative to number of words" , spellcheck          #8
			
			caps = round(capsCount(sentence.split()),2)                
			print "capscount relative to number of words", caps        #9

			punctu = round(punctCount(sentence),2)
			print "crazy punctuation relative to number of sentences" , punctu   #10
			

			print "............................................examples.............................................."


			digits = float(countDigits(sentence))
			print "digits", digits            #11

			percentC = float(percentCount(sentence))
			print "%", percentC          #12

			hyper = containsHyperlink(sentence)

			if average_hyperlink > 0.0:
				hyperlink_p = hyper     
			else:
				hyperlink_p = 0.0
			
			print "hyperlinks", hyperlink_p              #13                                                                  
			
			nee = preprocess_for_nee_and_print_ind(sentence)                
			#print "named entities mentioned in argument", nee

			if av_ne > 0.0:
				named_e = round((float(len(nee)) / av_ne),2)
			else:
				named_e = 0.0                                             
			
			print "named entities compared to debate ", named_e     #14                             
			  

			print ".........................................POS TAGS.........................................."

			#nouns in argument
			nouns_a = nouns_in_argument(token_sentence)                                   
			nr_nouns = round((float(len(nouns_a)) / words_argument),2)
			print " nouns%", nr_nouns                       #15

			adj_a = count_adjectives(sentence.split())
			nr_adj = round((float(len(adj_a)) / words_argument),2)
			print "adj%", nr_adj                                            #16

			verbs_a = count_verbs(sentence.split())
			nr_verbs = round((float(len(verbs_a)) / words_argument),2)
			print "verbs%", nr_verbs                             #17

			adverbs = count_adverbs(sentence.split())
			nr_adverbs = round((float(len(adverbs)) / words_argument),2)
			print "adverbs%", nr_adverbs                         #18,19

			prons = count_pronouns(sentence.split())
			nr_prons = round((float(len(prons)) / words_argument),2)
			print "pronons%", nr_prons                         #20

			print ".............................................long & rare words ............................................"

			_long_words = long_words(token_sentence)           #21
			#print "long words in argument", _long_words
			nr_long = round((float(len(_long_words)) / words_argument),2)
			print "longwords%", nr_long

			inter_rare = set(lemma_sent).intersection(very_rare)
			percent_rare = round(float(len(inter_rare)) / len(very_rare), 2)     #22
			print "intersection rare words", percent_rare

			average_freq = round(average_freqdist(lemma_sent, no_stop_lemmatised),2)       #23
			print "rarity of word, the lower the better", average_freq


			
			print ".............................................intersections with stuff........................................"
			#intersection of nouns in argument and most common nouns in debate
			inter_noun = set(nouns).intersection(nouns_a)
			percent_inter = round((float(len(inter_noun)) / len(nouns)),2)
			print "percent intersection nouns", percent_inter               #24


			#want to count the percentage of most common nouns in sentence
			noun_counter = 0
			for noun in set(nouns_a):
				if noun in nouns:
					noun_counter = noun_counter + nouns_a.count(noun)

			if noun_counter > 0:

				nounpercentage = round(float(noun_counter) / len(nouns_a),2)
				print "nounpercentace", nounpercentage                          #25
			else:
				nounpercentage = 0.0
				print "nounpercentage 0.0"


			
			#intersection of lemmas - see above
			inter_lemmas = set(lemma_sent).intersection(mc_lemma)
			percent_interII = round((float(len(inter_lemmas)) / len(mc_lemma)),2)
			print "percent intersection lemmas", percent_interII        #26

			
			lemma_counter = 0
			for lem in set(lemma_sent):
				if lem in mc_lemma:
					lemma_counter = lemma_counter + lemma_sent.count(lem)

			if lemma_counter > 0:

				lemmapercentage = round(float(lemma_counter) / len(lemma_sent),2)
				print "lemmapercentage", lemmapercentage
			else:
				lemmapercentage = 0.0
				print "lemmapercentage 0.0"                          #27
			
			#intersection of stems
			inter_stem = set(stem_sent).intersection(mc_stem)
			percent_interIII = round((float(len(inter_stem)) / len(mc_stem)),2)
			print "percent intersection stems", percent_interIII                #28

			


			stem_counter = 0
			for stem in set(stem_sent):
				#print stem
				if stem in mc_stem:
					#print stem
					stem_counter = stem_counter + stem_sent.count(stem)

			if stem_counter > 0:

				stempercentage = round(float(stem_counter) / len(stem_sent),2)
				print "stempercentage", stempercentage
			else:
				stempercentage = 0.0
				print "stempercentage 0.0"                      #29




			arg_bigrams = split_bigrams(all_lower_no_punct)
			inter_bigrams = set(bigrams).intersection(arg_bigrams)
			#print "intersectioned bigrams", inter_bigrams
			
			if len(bigrams) >0:
				percent_bigrams = round(float(len(inter_bigrams))/len(bigrams),2)
				print "percent intersec bigrams", percent_bigrams
			else:
				percent_bigrams = 0.0
				print "intersex bigrams: 0.0"                       #30

			
			bigram_counter = 0
			for bigram in set(arg_bigrams):
				if bigram in bigrams:
					bigram_counter = bigram_counter + arg_bigrams.count(bigram)

			if bigram_counter > 0:

				bigrampercentage = round(float(bigram_counter) / len(arg_bigrams),2)
				print "bigrampercentage", bigrampercentage
			else:
				bigrampercentage = 0.0
				print "bigrampercentage 0.0"                  #31




			arg_trigrams = split_trigrams(all_lower_no_punct)
			inter_trigrams = set(trigrams).intersection(arg_trigrams)
			#print "intersectioned trigrams", inter_trigrams

			if len(trigrams) > 0:
				percent_trigrams = round(float(len(inter_trigrams))/len(trigrams),2)
				print "percent intersec trigrams", percent_trigrams
			else: 
				percent_trigrams = 0.0
				print "percent intersec trigrams: 0.0"           #32


			trigram_counter = 0
			for trigram in set(arg_trigrams):
				if trigram in trigrams:
					trigram_counter = trigram_counter + arg_trigrams.count(trigram)

			if bigram_counter > 0:

				trigrampercentage = round(float(trigram_counter) / len(arg_trigrams),2)
				print "trigrampercentage", trigrampercentage
			else:
				trigrampercentage = 0.0
				print "trigrampercentage 0.0"               #33




			#"uncommon" words according googles most used words
			inter_google1 = set(unusual_wordsI).intersection(stem_sent)
			#print "intersectioned words google1", inter_google1
			percemt_g1 = round(float(len(inter_google1)) /len(unusual_wordsI),2)
			print "percent intersection unusual words google1", percemt_g1      #34

			inter_google2 = set(unusual_wordsII).intersection(stem_sent)
			#print "intersectioned words google1", inter_google2
			percemt_g2 = round(float(len(inter_google2)) /len(unusual_wordsII),2) 
			print "percent intersection unusual words google2", percemt_g2           #35
			print "***************************************************************************************************************+++"	
			###########################################################################################################################################



			#now we want to diferent vectors to make altering easier - one with statis stuff and one wih individual debate stuff (the intersections and most common)

			#VECTOR STRING
			vector_string_static = str(av_word_sen) + "," + str(av_len_word) + "," + str(word_ratio) + "," + str(sent_ratio) + "," + str(ratio_ws) + "," + str(linkwords) + "," + str(readability) + "," + str(badword) + "," + str(spellcheck) + "," + str(caps) + "," + str(punctu)  + "," + str(digits) + "," + str(percentC) + "," + str(hyperlink_p)  + "," + str(named_e) + "," + str(nr_nouns) + "," + str(nr_adj) + "," + str(nr_verbs) + "," + str(nr_adverbs) + "," + str(nr_adverbs) + "," +str(nr_prons) + "," + str(nr_long)  + "," + str(percent_rare)  + "," + str(average_freq) + "," + str(percent_inter) + "," + str(nounpercentage)  + "," + str(percent_interII) + "," + str(lemmapercentage) + "," + str(percent_interIII) + "," + str(stempercentage) + "," + str(percent_bigrams) + "," + str(bigrampercentage) + "," + str(percent_trigrams) + "," + str(trigrampercentage) + "," + str(percemt_g2) + "," + str(percemt_g2)
			print "VECTORSTRING"                                                                                                                                         #7                                                                                                                                                                                               
			print vector_string_static
			sheet.cell(row = i, column = 4).value = vector_string_static
			print "commas", vector_string_static.count(',')
			print "********************************************************************************************************************"
		wb.save(whole_name)


	print "done"


if __name__ == '__main__':
    demo()
