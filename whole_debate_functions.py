#clean and well documented code (hopefully)

#FOR WHOLE DEBATE CODE
import openpyxl
import os
import nltk,re
from not_ne_list import *
from ten_google_stem_list import*
from twen_google_stem_list import*
#from nltk.tokenize import word_tokenize 
from nltk.probability import FreqDist
from nltk.stem.wordnet import WordNetLemmatizer
from nltk.stem.porter import PorterStemmer
from nltk.stem import SnowballStemmer
from nltk.tag import pos_tag
from collections import Counter
from pattern.en import lemma, singularize
from enchant.checker import SpellChecker
from nltk.tokenize import word_tokenize, sent_tokenize

#for report: write about different results of stemming and lemmatising and which algorithm is better and which one is more efficient

#Probmem: one spreadsheet not working need to figure out why


###########################################################      PREPROCESSING        ##########################################################################
#preprocessing debate:
def preprocess_debateI(debate):

	#setting everything to lower to make processing easier
	debate = debate.lower() 

	#finding and counting hyperlinks, then delete them from debate (otherwise influence data)
	links = re.findall(r'(w?w?w?.?https?://[^\s]+)', debate)

	if len(links) != 0:
		for link in links:
			debate = debate.replace(link, '')
	
	#deleting all punctuation apart from apostrophes
	debate = re.sub("[^\w\d'\s]+", " ", debate)
	debate = debate.split()
	return debate  


def preprocess_debateII(debate):
	
	#removing stopwords
	with open('SmartStoplist.txt', 'r') as myfile:
		data=myfile.read().replace('\n', ' ')
	
	data=data.split()

	meaningful_words = [w for w in debate if not w in data and w.isalpha()]
	#print meaningful_words
	return meaningful_words   

#########################################################          MOST COMMON STUFF          ################################################################
# most common words, might not use that
def most_common_meaningful(debate):

	#frequency distribution
	fdist_normal = FreqDist(debate)
	list_most_common_words = fdist_normal.most_common(20)

	meaningful_list = []
	for list_tuple in list_most_common_words:
		meaningful_list.append(list_tuple[0])

	return meaningful_list

#most common lemmas
#using NLTK lemmatiser
"""
def lemmatised_words(debate):	
	
	lmtsr = WordNetLemmatizer()
	#meaningful_string = word_tokenize(meaningful_string)
	lemmatised_words = [lmtsr.lemmatize(w) for w in debate]
	#print lemmatised_words

	fdist = FreqDist(lemmatised_words)

	list_mc = fdist.most_common(20)

	most_common_lemmatised_list = []
	
	for list_tuple in list_mc:
		most_common_lemmatised_list.append(list_tuple[0].encode('utf-8'))

	return most_common_lemmatised_list

"""	
#using pattern lemmatiser (in my opinion better)
def lemmatised_words2(debate):	
	
	lemmatised_words = [lemma(w) for w in debate]
	#print lemmatised_words

	fdist = FreqDist(lemmatised_words)

	list_mc = fdist.most_common(20)

	most_common_lemmatised_list = []
	
	for list_tuple in list_mc:
		most_common_lemmatised_list.append(list_tuple[0].encode('utf-8'))

	return most_common_lemmatised_list

#stemming using snowball stemmer
def stemmed_snowball(debate):

	snowball = SnowballStemmer("english")

	stemmed_words_snowball = [snowball.stem(w) for w in debate]

	fdist2 = FreqDist(stemmed_words_snowball)
	list_mc2 = fdist2.most_common(20)

	most_common_snowball = []
	
	for list_tuple in list_mc2:

		most_common_snowball.append(list_tuple[0].encode('utf-8'))

	return most_common_snowball
	######################################################################
"""
def stemmed_porter(debate):
	porter = PorterStemmer()
	stemmed_words_porter = [porter.stem(w) for w in debate]

	fdist3 = FreqDist(stemmed_words_porter)

	list_mc3 = fdist3.most_common(20)

	most_common_porter = []
	
	for list_tuple in list_mc3:
		most_common_porter.append(list_tuple[0].encode('utf-8'))

	return most_common_porter
"""		

########################################################## most common nouns ##############################################################################

def extract_nouns(debate):
	#print debate
	
	tagged = pos_tag(debate)
	
	propernouns = [word for word,pos in tagged if pos == 'NNP']  #should be able to make or statement no?
	nnouns = [word for word,pos in tagged if pos == 'NN']
	nounss = [word for word,pos in tagged if pos == 'NNS']
	
	nouns = nnouns + propernouns + nounss

	c = Counter(nouns)
	most_common_nouns = c.most_common(20)
	mcn_list = []
	for list_tuple in most_common_nouns:
		mcn_list.append(list_tuple[0])
	
	return mcn_list



########################################################## UNUSUAL WORDS #####################################################################################

def unusual_words(text, mcs, google):    
    text = text.split()

    #take text
    # spellcheck it
    # put correct words in list
    # stem them
    #check if word stem in one of the lists
    #return list of stems not in google list
    spellC = SpellChecker("en_GB")
    correct_words = []

    for w in text:
    	if w.isalpha():
    		correct = spellC.check(w)
    		if correct == True:
    			correct_words.append(w.encode('utf-8'))

    snowball = SnowballStemmer("english")

    delete_list = mcs + google

    stemmed_words = [snowball.stem(w) for w in correct_words]
    unusual_words = [w.encode('utf-8') for w in stemmed_words if w not in delete_list]
    
    return unusual_words                    #will miss words that are not in the enchant checker dictionary!!!!!!       


def preprocess_for_nee_and_print(argument, mcw, mcl):                        #need to write better function that might correct the sentence first before extracting them but for now it will do it
	

	sentences = nltk.sent_tokenize(argument)
	tokenized_sentences = [nltk.word_tokenize(sentence) for sentence in sentences]
	tagged_sentences = [nltk.pos_tag(sentence) for sentence in tokenized_sentences]
	chunked_sentences = nltk.ne_chunk_sents(tagged_sentences, binary=True)

	entity_names = []

	for tree in chunked_sentences:
		#tree.draw()
		entity_names.extend(extract_entity_names(tree))

	#print entity_names
	mcwl = mcw + mcl
	entities = [w for w in entity_names if w not in not_NE]
	#print entities
	named_entities = [w for w in entities if w.lower() not in mcwl]
	named_entities = set(named_entities)
	print named_entities

def extract_entity_names(tree):
    entity_names = []

    if hasattr(tree, 'label') and tree.label:
        if tree.label() == 'NE':
            entity_names.append(' '.join([child[0] for child in tree]))
        else:
            for child in tree:
                entity_names.extend(extract_entity_names(child))

    return entity_names





#arg1 = '''It goes against human nature and, as www.http://dasjd.dkjsad other fellow voters have mentioned, the definition of "Marriage". And you can't exclude religion from this argument, because thats what happened when people don't have faith or belief in any religion. Religion is there to act as guidelines to what is right and what is wrong. Religion is a way of life. Without it, we're lost.'''


directory = '/home/user/Documents/IndProject/Data_and_Code/INDIVIDUAL_ARGUMENTS/'

"""
wb = openpyxl.load_workbook('/home/user/Documents/IndProject/Data_and_Code/INDIVIDUAL_ARGUMENTS/india_no.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
end = sheet.max_row
text = []

for i in range(2,end,1):
	sentence = [sheet.cell(row = i, column =2).value.encode('utf-8')]
	text = text+sentence
	#print type(sentence)

text= ' '.join(str(r) for r in text)

prep = preprocess_debateI(text)

preprocessed = preprocess_debateII(prep)
nouns = extract_nouns(preprocessed)

mcw_full = most_common_meaningful(preprocessed)
#mcw_lemmatised = lemmatised_words(preprocessed)
mcw_lemmatised2 = lemmatised_words2(preprocessed)
mcw_stemmed_snowball = stemmed_snowball(preprocessed)
#mcw_stemmed_porter = stemmed_porter(preprocessed)

print ""
print "most common words", mcw_full
#print "most common lemma nltk:", mcw_lemmatised  #probably most useful!!
print "most common lemma pattern:" , mcw_lemmatised2
print "stemmed snowball", mcw_stemmed_snowball
#print "setemmed porter", mcw_stemmed_porter
print "nouns", nouns
unusualwords = unusual_words(text, mcw_lemmatised2)
print "unusual", unusualwords

least_common(preprocessed)



counter = 0
for file in os.listdir(directory):
	print "******************************************************************************************************************************"
	print "name of file:" , file
	whole_name = directory+file
	wb = openpyxl.load_workbook(whole_name)
	sheet = wb.get_sheet_by_name('Sheet1')
	end = sheet.max_row
	text = []
	

	for i in range(2,end,1):
		sentence = sheet.cell(row = i, column =2).value.encode('utf-8')
		#print sentence

		prep1 = preprocess_debateI(sentence)
		prep2 = preprocess_debateII(prep1)
		nouns = extract_nouns(prep2)
		mcw_full = most_common_meaningful(prep2)
		#mcw_lemmatised = lemmatised_words(preprocessed)
		mcw_lemmatised2 = lemmatised_words2(prep2)
		mcw_stemmed_snowball = stemmed_snowball(prep2)

		print prep1
		print prep2
		print nouns
		print mcw_full
		print mcw_lemmatised2
		print mcw_stemmed_snowball
	
	counter +=1


print counter
"""
for file in os.listdir(directory):
	print "******************************************************************************************************************************"
	print "name of file:" , file
	whole_name = directory+file
	wb = openpyxl.load_workbook(whole_name)
	sheet = wb.get_sheet_by_name('Sheet1')
	end = sheet.max_row
	text = []

	for i in range(2,end,1):
		sentence = [sheet.cell(row = i, column =2).value.encode('utf-8')]
		text = text+sentence
	#print type(sentence)

	text= ' '.join(str(r) for r in text)
	#print text

	
	prep = preprocess_debateI(text)
	preprocessed = preprocess_debateII(prep)
	nouns = extract_nouns(preprocessed)

	mcw_full = most_common_meaningful(preprocessed)
	#mcw_lemmatised = lemmatised_words(preprocessed)
	mcw_lemmatised2 = lemmatised_words2(preprocessed)
	mcw_stemmed_snowball = stemmed_snowball(preprocessed)
	#mcw_stemmed_porter = stemmed_porter(preprocessed)

	print ""
	print "most common words", mcw_full
	#print "most common lemma nltk:", mcw_lemmatised  #probably most useful!!
	print "most common lemma pattern:" , mcw_lemmatised2
	print "stemmed snowball", mcw_stemmed_snowball
	#print "setemmed porter", mcw_stemmed_porter
	print "nouns", nouns
	
	unusualstems = unusual_words(text, mcw_stemmed_snowball, google10_stemed)
	print "unusual", unusualstems

	ne = preprocess_for_nee_and_print(text, mcw_lemmatised2, mcw_full)
	print ne
	print "******************************************************************************************************************************"

#wb = openpyxl.load_workbook('/home/user/Documents/IndProject/Data_and_Code/INDIVIDUAL_ARGUMENTS/abortion_contra.xlsx')
#sheet = wb.get_sheet_by_name('Sheet1')
