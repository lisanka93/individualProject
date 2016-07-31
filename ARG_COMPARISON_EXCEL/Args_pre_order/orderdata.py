# -*- coding: utf-8 -*-

import openpyxl
import os


directory = '/home/user/Documents/IndProject/Data_and_Code/ARG_COMPARISON_EXCEL/ARGS/'

for file in os.listdir(directory):
	print file
	file_name = directory+file
	wb = openpyxl.load_workbook(file_name)
	sheet = wb.get_sheet_by_name('Sheet1')
	end = sheet.max_row

	#print sheet.cell(row = 3, column = 5).value 

	for i in range(2,end+1,1):           #start in second row until end in one steps
		args = sheet.cell(row = i, column =1).value
		args = args.split()
		id1 = args[0]
		id2 = args[1]
		value = sheet.cell(row = i, column = 2).value
		#print type(value)
		#print value
		
		if value == "a1":
			
			value = "1,0"
			sheet.cell(row = i, column = 3).value = value
		else:
			value = "0,1"
			sheet.cell(row = i, column = 3).value = value
		
		sheet.cell(row = i, column = 1).value = id1
		sheet.cell(row = i, column = 2).value = id2

	wb.save(file_name)