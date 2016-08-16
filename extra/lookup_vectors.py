# -*- coding: utf-8 -*-

import openpyxl
import os
import numpy


directory1 = '/home/user/Documents/IndProject/Data_and_Code/test_excel/one/'

directory2 = '/home/user/Documents/IndProject/Data_and_Code/test_excel/two/'


for file in os.listdir(directory1):
	#print file

	file_name1 = directory1+file
	file_name2 = directory2+file
	wb1 = openpyxl.load_workbook(file_name1)
	wb2 = openpyxl.load_workbook(file_name2)
	sheet1 = wb1.get_sheet_by_name('Sheet1')
	sheet2 = wb2.get_sheet_by_name('Sheet1')

	end = sheet1.max_row
	end2 = sheet2.max_row

	#print sheet.cell(row = 3, column = 5).value 

	for i in range(2,end+1,1):           #start in second row until end in one steps
		id1 = sheet1.cell(row = i, column =1).value
		id2 = sheet1.cell(row = i, column =2).value
		
		for j in range(2,end2+1,1):
			value = sheet2.cell(row = j, column =1).value
			if value == id1:
				vector1 = sheet2.cell(row = j, column =3).value
				#print vector1

		for k in range(2,end2+1,1):
			value = sheet2.cell(row = k, column =1).value
			if value == id2:
				vector2 = sheet2.cell(row = k, column =3).value


		vector = vector1 + "," + vector2
		#print vector

		sheet1.cell(row = i, column = 4).value = vector
		vector = [vector.encode('utf-8')]

		#print type(value)
		#print value

	wb1.save(file_name1)


