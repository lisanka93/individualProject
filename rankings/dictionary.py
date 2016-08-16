# -*- coding: utf-8 -*-

import openpyxl
import os
from collections import Counter


#the two directories with the 2 different spreadsheets

#this one for the comparison -> vector will be dumped into the 4th column!
directory = '/home/user/Documents/IndProject/Data_and_Code/dict/'
print "hello"

for file in os.listdir(directory):
	print 'hello'

	file_name = directory+file
	wb = openpyxl.load_workbook(file_name)
	print file_name
	sheet = wb.get_sheet_by_name('Sheet1')
	end = sheet.max_row
	collection = []

	for i in range(1, end+1, 1):
		if sheet.cell(row = i, column = 3).value == '1,0':
			#sheet.cell(row = i, column = 4).value = sheet.cell(row = i, column = 1).value
			collection.append(sheet.cell(row = i, column = 1).value)
		else:
			#sheet.cell(row = i, column = 4).value = sheet.cell(row = i, column = 2).value
			collection.append(sheet.cell(row = i, column = 2).value)


	wb.save(file_name)

	print collection
	occurances = Counter(collection)
	print occurances
	#occurances = list(occurances)
	#print occurances


	print "Rankings:"

	for arg, rank in occurances.items():
		print (arg, rank)
		#print "argument: ", arg[0], " count: ", arg[1]





