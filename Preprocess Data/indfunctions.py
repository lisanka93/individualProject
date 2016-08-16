
#PROCESS DATA
import openpyxl
import os
from whole_debate_func_clean import *
from ind_arguments import*
#import nltk
#from nltk.stem import SnowballStemmer
#**********************************************************************************************************************
"""right so this loops through the individual argument folder and analyses the debate and the individual arguments"""
#**********************************************************************************************************************
def demo():
	directory = '/home/user/Documents/IndProject/Data_and_Code/INDIVIDUAL_ARGUMENTS/'
	for file in os.listdir(directory):
		print "******************************************************************************************************************************"
		print "name of file:" , file
		argument_counter = 0

		whole_name = directory+file
		wb = openpyxl.load_workbook(whole_name)
		sheet = wb.get_sheet_by_name('Sheet1')
		end = sheet.max_row
		text = []
		#let the fun begin: first append all sentences to one big text
		for i in range(2,end,1):
			sentence = [sheet.cell(row = i, column =2).value.encode('utf-8')]
			text = text+sentence
			argument_counter += 1
		#print type(sentence)
		text= ' '.join(str(r) for r in text)
		#print text                              #great, works!
		text_length = len(text.split())                                                     #***august: check !
		print "text length", text_length
		print "that many arguments", argument_counter

		debate_sentences = len(nltk.sent_tokenize(text))
		print "that many sentences", debate_sentences
		###############################################################################################
		#average length of argument

		average_length = ((float(text_length)) / argument_counter ) 
		print "average length of arguments in words", average_length

		average_sent = ((float(debate_sentences)) / argument_counter)
		print "average length of arguments in sentences", average_sent

		average_sentlength = ((float(text_length)) / debate_sentences)
		print "average sent length", average_sentlength

		tuple_hyp = preprocess_debateI(text)
		
		#number of hyperlinks mentioned in whole debate
		number_hyperlinks = tuple_hyp[1]
		print "hyperlinks in whole debate: ", number_hyperlinks
		
		average_hyperlink = (float(number_hyperlinks) / argument_counter)
		print "average number of hyperlinks per argument", average_hyperlink                                                  #***august: check!
		
		#no_punct is the whole debate without punctuation and tokenized in separate words
		no_punct = tuple_hyp[0]
		
		#all meaningful/relevant words
		no_stopwords = preprocess_debateII(no_punct)

		no_stop_lemmatised = preprocess_debateIII(no_stopwords)

		#long words
		longWords = long_words(no_stop_lemmatised)
		print "long words", longWords

		#very rare
		very_rare  = very_rare_long_words(no_stop_lemmatised, longWords)
		print "very rare", very_rare
		
		#most common lemmas and stems
		mc_lemma = lemmatised_words(no_stop_lemmatised)
		print "mc lemmas", mc_lemma
		
		mc_stem = stemmed_snowball(no_stopwords)
		print "mc stems", mc_stem
		
		#most common nouns
		nouns = extract_nouns(no_stopwords)
		print "mc nouns", nouns

		bigram_extr(no_punct)
		trigram_extr(no_punct)


		#named entities
		named_entities = preprocess_for_nee_and_print(text)
		print "named entities in whole debate", named_entities    
		
		#print "HERE" ,named_entities                                           #august: check SHOULD RETURN LIST NOT SET
		av_ne = round((float(len(named_entities)) / argument_counter),2)
		print "average number of NE mentioned in argument", av_ne                                                                  #august CHECK!! average ne per argument
		
		#unusual words compared to most common 10000 google words
		print "UNUSUAL WORDS ACCORDING TO MOST COMMONLY USED GOOGLE"
		
		unusual_wordsI = unusual_words(text, mc_stem, google10_stemed)
		print "10000 google words", unusual_wordsI
		
		#unusual words compared to most common 20000 google words
		unusual_wordsII = unusual_words(text, mc_stem, google20_stemed)
		print "20000 google words" ,unusual_wordsII

		# UP TO HERE EVERYTHING SEEMS RIGHT YEAY


		#######################################################################################################################
		
		print "_____________________________ individual arguments ____________________________________________________________"
		print "results sentence"
		#prints all the interesting values of each individual argument
		for i in range(2, end+1, 1):

			#sentence aka argument in one cell
			sentence = sheet.cell(row = i, column =2).value.encode('utf-8')
			argu_length = len(sentence)
			print "argulength", argu_length

			#################################################################################
			
			#super preprocessed for mc stuff extraction
			token_sentence = preprocess_argumentIV(sentence)         
			
			#string without hyperlinks
			no_hyperlinks= preprocess_argumentI(sentence)
			
			#delete all punctuatuin apart from apostrophe (also string) 
			only_apostrophe_alllower = preprocess_argumentII(no_hyperlinks)

			#list
			all_lower_no_punct = preprocess_argumentIII(only_apostrophe_alllower)

			#lemmatised sentence
			lemma_sent = [lemma(w).encode('utf-8') for w in token_sentence]
			
			#stemmed sentence
			snowball = SnowballStemmer("english")
			
			stem_sent  = [snowball.stem(w).encode('utf-8') for w in token_sentence]
			
			##################################################################################
			

			nr_sentences = len(nltk.sent_tokenize(sentence))
			print "number of sentences in argument", nr_sentences

			words_argument = len(sentence.split())
			print "words in argument", words_argument

			av_word_sen = round((float(words_argument)/nr_sentences),2)
			print "averagde number of words per sentence in argument", av_word_sen

			av_len_word = round((float(argu_length) / words_argument),2)
			print "average length of word", av_len_word

			#argument longer than average (words)?
			word_ratio = round(float(words_argument) / average_length,2)
			print "word ratio", word_ratio


			#argument longer than average (sentence)?
			sent_ratio = round(float(nr_sentences) / average_sent,2)
			print "sent rati0", sent_ratio


			#sentence lenth longer than average?
			ratio_ws = round(float(av_word_sen) / average_sentlength,2)
			print "ratio ws", ratio_ws

			
			linkwords = linkingwordcount(all_lower_no_punct)                  #change formula - linking words compared to sentences not words in argument!
			linkwords = round((float(linkwords) / nr_sentences),2)
			print " linnking words relative to sentence", linkwords
			
			readability = round(ColemanLiauIndex(no_hyperlinks),2)
			print "readability", readability
			
			badword = round(badwordcount(all_lower_no_punct),2)                       #august: maybe just change ti number
			print "badwords",  badword
			
			spellcheck = round(spellCheck(only_apostrophe_alllower.split()),2)
			print "errors relative to number of words" , spellcheck           
			

			punctu = round(punctCount(sentence),2)
			print "crazy punctuation relative to number of sentences" , punctu
			
			digits = countDigits(sentence)
			print "digits", digits

			percentC = percentCount(sentence)
			print "%", percentC

			hyper = containsHyperlink(sentence)

			if average_hyperlink >= 1:
				hyperlink_p = hyper      #change!!!!!!!!
			else:
				hyperlink_p = 0.0
			
			print "hyperlinks", hyperlink_p                                                                           #******august test"!!!!!
			
			caps = round(capsCount(sentence.split()),2)                 #need all words
			print "capscount relative to number of words", caps
			
			nee = preprocess_for_nee_and_print_ind(sentence)                 #need list for average, not set
			print "named entities mentioned in argument", nee

			if av_ne > 0.0:
				named_e = round((float(len(nee)) / av_ne),2)
			else:
				named_e = 0.0                                             
			
			print "named entities compared to debate ", named_e                                                  #august: check!!!
			
			#nouns in argument
			nouns_a = nouns_in_argument(token_sentence)                 #without stopwords!!                      
			print "nouns in argument", nouns_a
			nr_nouns = round((float(len(nouns_a)) / words_argument),2)
			print " nouns%", nr_nouns

			adj_a = count_adjectives(sentence.split())
			print "adjectives in arugument", adj_a
			nr_adj = round((float(len(adj_a)) / words_argument),2)
			print "adj%", nr_adj

			verbs_a = count_verbs(sentence.split())
			print "verbs in argument", verbs_a
			nr_verbs = round((float(len(verbs_a)) / words_argument),2)
			print "verbs%", nr_verbs

			adverbs = count_adverbs(sentence.split())
			print "adverbs in argument", adverbs
			nr_adverbs = round((float(len(adverbs)) / words_argument),2)
			print "adverbs%", nr_adverbs

			prons = count_pronouns(sentence.split())
			print "pronouns in argument", prons
			nr_prons = round((float(len(prons)) / words_argument),2)
			print "pronons%", nr_prons

			_long_words = long_words(token_sentence)
			print "long words in argument", _long_words
			nr_long = round((float(len(_long_words)) / words_argument),2)
			print "longwords%", nr_long

			inter_rare = set(lemma_sent).intersection(very_rare)
			percent_rare = round(float(len(inter_rare)) / len(very_rare), 2)
			print "intersection rare words", percent_rare

			average_freq = round(average_freqdist(lemma_sent, no_stop_lemmatised),2)
			print "rarity of word, the lower the better", average_freq

			#######################################################################################################################

			
			#intersection of nouns in argument and most common nouns in debate
			inter_noun = set(nouns).intersection(nouns_a)
			percent_inter = round((float(len(inter_noun)) / len(nouns)),2)
			print "percent intersection nouns", percent_inter
			
			#intersection of lemmas - see above
			inter_lemmas = set(lemma_sent).intersection(mc_lemma)
			percent_interII = round((float(len(inter_lemmas)) / len(mc_lemma)),2)
			print "percent intersection lemmas", percent_interII
			
			#intersection of stems
			inter_stem = set(stem_sent).intersection(mc_stem)
			percent_interIII = round((float(len(inter_stem)) / len(mc_stem)),2)
			print "percent intersection stems", percent_interIII
			
			#"uncommon" words according googles most used words
			inter_google1 = set(unusual_wordsI).intersection(stem_sent)
			print "intersectioned words google1", inter_google1
			percemt_g1 = round(float(len(inter_google1)) /len(unusual_wordsI),2)
			print "percent intersection unusual words google1", percemt_g1

			inter_google2 = set(unusual_wordsII).intersection(stem_sent)
			print "intersectioned words google1", inter_google2
			percemt_g2 = round(float(len(inter_google2)) /len(unusual_wordsII),2) 
			print "percent intersection unusual words google2", percemt_g2           #
			print "***************************************************************************************************************+++"	
			###########################################################################################################################################



			#now we want to diferent vectors to make altering easier - one with statis stuff and one wih individual debate stuff (the intersections and most common)

			#VECTOR STRING
			vector_string_static = str(av_word_sen) + "," + str(av_len_word) + "," + str(word_ratio) + "," + str(sent_ratio) + "," + str(ratio_ws) + "," + str(linkwords) + "," + str(readability) + "," + str(badword) + "," +str(spellcheck) + "," + str(punctu)  + "," + str(digits) + "," + str(percentC) + "," + str(hyperlink_p) + "," + str(caps) + "," + str(named_e) + "," + str(nr_nouns) + "," + str(nr_adj) + "," + str(nr_verbs) + "," + str(nr_adverbs) + "," + str(nr_adverbs) + "," +str(nr_prons) + "," + str(nr_long)  + "," + str(percent_rare) + "," + str(percemt_g2) + "," + str(percemt_g2) + "," + str(average_freq)
			print "VECTORSTRING"                                                                                                                                         #7                                                                                                                                                                                               
			print vector_string_static
			sheet.cell(row = i, column = 3).value = vector_string_static
			print "********************************************************************************************************************"
		wb.save(whole_name)


	#wb.save(whole_name)
	print "done"




		##### OK SEEMS TO WORK











	#print counter
if __name__ == '__main__':
    demo()