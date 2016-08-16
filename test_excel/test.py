
import openpyxl
import os




directory = '/home/user/Documents/IndProject/Data_and_Code/ARG_COMPARISON_EXCEL/ARGS_post_order/'

#############################################################################################################
directory_xy = '/home/user/Documents/IndProject/Data_and_Code/test_excel/x_y/'

filename_x = 'x.xlsx'
filename_y = 'y.xlsx'

x_file = directory_xy + filename_x
y_file = directory_xy + filename_y

wb_x = openpyxl.load_workbook(x_file)
wb_y = openpyxl.load_workbook(y_file)
x_sheet = wb_x.get_sheet_by_name('Sheet1')
y_sheet = wb_y.get_sheet_by_name('Sheet1')
###############################################################################################################
crawler = 1

for file in os.listdir(directory):
	#print file

	file_name = directory+file
	wb = openpyxl.load_workbook(file_name)
	sheet1 = wb.get_sheet_by_name('Sheet1')

	end = sheet1.max_row
	x_sheet.cell(row = crawler, column = 1).value = file
	y_sheet.cell(row = crawler, column = 1).value = file

	for i in range(2,end+1,1):
		vector = sheet1.cell(row = i, column =4).value      #vector from column 4 
		better = sheet1.cell(row = i, column =3).value      #comparison from column 3
		x_sheet.cell(row = (i+crawler-1), column = 1).value = vector                        #vectorcomes into x sheet - input
		#print vector
		y_sheet.cell(row = (i+crawler-1), column = 1).value = better                        #comparison into y sheet  -output
		#print better

	crawler = crawler+end

print "done"
wb_x.save(x_file)
wb_y.save(y_file)





