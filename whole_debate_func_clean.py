"""
#PROCESS DATA

import openpyxl
import os
from whole_debate_func_clean import *
from ind_arguments import*


#**********************************************************************************************************************
"""right so this loops through the individual argument folder and analyses the debate and the individual arguments"""
#**********************************************************************************************************************

def demo():

	directory = '/home/user/Documents/IndProject/Data_and_Code/INDIVIDUAL_ARGUMENTS/'

	for file in os.listdir(directory):
		print "******************************************************************************************************************************"
		print "name of file:" , file
		argument_counter = 0
		whole_name = directory+file
		wb = openpyxl.load_workbook(whole_name)
		sheet = wb.get_sheet_by_name('Sheet1')
		end = sheet.max_row
		text = []

		#let the fun begin: first append all sentences to one big text
		for i in range(2,end,1):
			sentence = [sheet.cell(row = i, column =2).value.encode('utf-8')]
			text = text+sentence
			argument_counter += 1
		#print type(sentence)

		text= ' '.join(str(r) for r in text)
		#print text                              #great, works!
		text_length = len(text.split())                                                     #***august: check !
		print text_length
		print argument_counter

		###############################################################################################

		#average length of argument
		average_length = (float(len(text.split()) / argument_counter)                         #*****august check!!

		print "results whole debate: "
		tuple_hyp = preprocess_debateI(text)

		#number of hyperlinks mentioned in whole debate
		number_hyperlinks = tuple_hyp[1]
		#print "hyperlinks: ", number_hyperlinks
		average_hyperlink = (float(number_hyperlinks) / argument_counter)
		print "average link", average_hyperlink                                                  #***august: check!
		
		#no_punct is the whole debate without punctuation and tokenized in separate words
		no_punct = tuple_hyp[0]
		
		#all meaningful/relevant words
		no_stopwords = preprocess_debateII(no_punct)
		
		#most common lemmas and stems
		mc_lemma = lemmatised_words(no_stopwords)
		print "mc lemmas", mc_lemma

		mc_stem = stemmed_snowball(no_stopwords)
		print "mc stems", mc_stem

		#most common nouns
		nouns = extract_nouns(no_stopwords)
		print "mc nouns", nouns

		#named entities
		named_entities = preprocess_for_nee_and_print(text)
		print "ne", named_entities                                               #august: check SHOULD RETURN LIST NOT SET
		av_ne = (float(named_entities) / argument_counter)
		print av_ne                                                                  #august CHECK!! average ne per argument

		#unusual words compared to most common 10000 google words
		print "google"
		unusual_wordsI = unusual_words(text, mc_stem, google10_stemed)
		print unusual_wordsI

		#unusual words compared to most common 20000 google words
		unusual_wordsII = unusual_words(text, mc_stem, google20_stemed)
		print unusual_wordsII
		#######################################################################################################################
		
		print "_____________________________ individual arguments ____________________________________________________________"
		print "results sentence"
		for i in range(2, end, 1):
			#sentence aka argument in one cell
			sentence = sheet.cell(row = i, column =2).value.encode('utf-8')
			#################################################################################
			#super preprocessed for mc stuff extraction
			token_sentence = preprocess_argumentIV(sentence)         #sentence without stopwords
			
			#string without hyperlinks
			no_hyperlinks= preprocess_argumentI(sentence)
			#delete all punctuatuin apart from apostrophe (also string) 
			only_apostrophe_alllower = preprocess_argumentII(no_hyperlinks)

			#list
			all_lower_no_punct = preprocess_argumentIII(only_apostrophe_alllower)

			lemma_sent = [lemma(w) for w in token_sentence]
			snowball = SnowballStemmer("english")
			stem_sent  = [snowball.stem(w) for w in token_sentence]
			##################################################################################
			
			#length


			#linkingwordcount
			linkwords = linkingwordcount(all_lower_no_punct)                  #change formula - linking words compared to sentences not words in argument!
			print " linnking", linkwords

			readability = ColemanLiauIndex(no_hyperlinks)
			print "read", readability

			badword = badwordcount(all_lower_no_punct)                       #august: maybe just change ti number
			print "badwords",  badword

			spellcheck = spellCheck(only_apostrophe_alllower.split())
			print "errors" , spellcheck           

			punctu = punctCount(sentence)
			print "punct" , punctu

			hyper = containsHyperlink(sentence)
			hypenlink_p = (float(hyper) / average_hyperlink)
			print hypenlink_p                                                                           #******august test"!!!!!

			caps = capsCount(sentence.split())                 #need all words
			print "caps", caps

			nee = preprocess_for_nee_and_print_ind(sentence)                 #need list for average, not set
			print "named entity", nee

			named_e = (float(len(nee)) / av_ne)                                                       #**august CHECK

			print ne_counter                                                   #august: check!!!

			nouns_a = nouns_in_argument(token_sentence)                 #without stopwords!!                      
			print "nouns", nouns_a

			#intersection of nouns
			inter_noun = set(nouns).intersection(nouns_a)
			percent_inter = (float(len(inter_noun)) / len(nouns))
			print "percent intersection nouns", percent_inter

			#intersection of lemmas
			inter_lemmas = set(lemma_sent).intersection(mc_lemma)
			percent_interII = (float(len(inter_lemmas)) / len(mc_lemma))
			print "percent intersection lemmas", percent_interII

			#intersection of stems
			inter_stem = set(stem_sent).intersection(mc_stem)
			percent_interIII = (float(len(inter_stem)) / len(mc_stem))
			print "percent intersection stems", percent_interIII

			#intersection of unusual words!	                                                                              #agusu5: cheek h3m 
			# here it goes
			stem_argument = [stemmed_snowball(w) for w in token_sentence]
			inter_google1 = set(stem_argument).intersection(token_sentence))
			percemt_g1 = float(len(inter_google1)) /len(unusual_wordsI)
			
			stem_argument2 = [stemmed_snowball(w) for w in token_sentence]
			inter_google2 = set(stem_argument2).intersection(token_sentence)
			percemt_g1 = float(len(inter_google2)) /len(unusual_wordsII)             #






		######################################################################################################

		#could do the same for lemma and see if result is better but ill leave it like that for now!


	#print counter



if __name__ == '__main__':
    demo()
"""



#whole debate functions clean

import nltk,re
from not_ne_list import *
from ten_google_stem_list import*
from twen_google_stem_list import*
#from nltk.tokenize import word_tokenize 
from nltk.probability import FreqDist
#from nltk.stem.wordnet import WordNetLemmatizer
from nltk.stem.porter import PorterStemmer
from nltk.stem import SnowballStemmer
from nltk.tag import pos_tag
from collections import Counter
from pattern.en import lemma, singularize
from enchant.checker import SpellChecker
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.collocations import *



###########################################################      PREPROCESSING      ##########################################################################
#preprocessing debate:

def preprocess_debateI(debate):
	"""sets everything to lower case and removes hyperlinks in order to avoid spelling mistakes. after finding
		and counting all hyperlinks also deletes all punctuation and returns tokanized debate / list"""

	debate = debate.lower() 

	links = re.findall(r'(w?w?w?.?https?://[^\s]+)', debate)

	x = len(links)

	if x != 0:
		for link in links:
			debate = debate.replace(link, '')

	debate = re.sub("[^\w\d'\s]+", "", debate)   
	debate = debate.split()
	return (debate, x)                                 


def preprocess_debateII(debate):
	"""removes all stopwords and returns list of most meaningful words in debate"""
	
	with open('SmartStoplist.txt', 'r') as myfile:
		data=myfile.read().replace('\n', ' ')
	
	data=data.split()

	meaningful_words = [w for w in debate if not w in data and w.isalpha()]

	return meaningful_words                         #returns list of debate words - stopwords



def preprocess_debateIII(debate):
	lemmatised_words = [lemma(w) for w in debate]

	return lemmatised_words



######################################## preprocessing indifivual arguments #####################################################################


def preprocess_argumentI(debate):
	"""sets everything to lower and also removes hyperlinks from argument - returns STRING"""
	
	debate = debate.lower() 

	links = re.findall(r'(w?w?w?.?https?://[^\s]+)', debate)

	x = len(links)

	if x != 0:
		for link in links:
			debate = debate.replace(link, '')
	
	return debate                  


def preprocess_argumentII(debate):
	#removes punctuation APART FROM APOSTROPHE and sets to lower
	
	debate = re.sub("[^\w\d'\s]+", "", debate)  
	debate = debate.lower()
	return debate


def preprocess_argumentIII(debate):
	#also removes apostrophe and splits argument - returns list

	debate = re.sub("'", "", debate)   
	debate = debate.split()
	return debate


def preprocess_argumentIV(debate):
	#all of the above plus no stopwords

	debate = debate.lower() 
	debate = re.sub("[^\w\d'\s]+", "", debate)  
	debate = debate.split()
	with open('SmartStoplist.txt', 'r') as myfile:
		data=myfile.read().replace('\n', ' ')
	
	data=data.split()

	meaningful_words = [w for w in debate if not w in data and w.isalpha()]
	
	return meaningful_words  


##############################################bigrams   & trigrams     #################################################
	
def bigram_extr(meaningful_words):             #nostopwords 
	bigram_measures = nltk.collocations.BigramAssocMeasures()
	finder = BigramCollocationFinder.from_words(meaningful_words)
	scores = finder.score_ngrams(bigram_measures.raw_freq)
	scores_25 = scores[:25]                                                          #only first five are accurate, otherwise biased
	scores_5 = scores[:5]

	#print "25 bigrams:", scores_25
	#print "5 bigrams: ",scores_5
	#print "bigrams", finder.nbest(bigram_measures.pmi, 100)

def trigram_extr(meaningful_words):
	trigram_measures = nltk.collocations.TrigramAssocMeasures()
	finderII = TrigramCollocationFinder.from_words(meaningful_words)
	scoresI = finderII.score_ngrams(trigram_measures.raw_freq)
	scoresI_25 = scoresI[:15]                                                          #only first five are accurate, otherwise biased
	scoresI_5 = scoresI[:5]

	#print "15 triigrams:", scoresI_25
	#print "5 triigrams: ", scoresI_5


######################################################## long words ####################################################################


def long_words(debate):                                    #same call as for lemmatised wods

	lemmatised_words = debate
	#lemmatised_words = [lemma(w) for w in debate]
	#print lemmatised_words

	fdist = FreqDist(lemmatised_words)
	#vocabulary = fdist.keys()

	long_words = [w for w in lemmatised_words if len(w) > 9]
	long_words = set(long_words)

	long_words_correct = []
	chkr = SpellChecker("en_GB", debate)

	for word in long_words:
		if chkr.check(word) == True:
			long_words_correct.append(word.encode('utf-8'))

	return long_words_correct             #returns list of long words


def very_rare_long_words(debate, longwords):             #takes lemmatised debate

	fdist = FreqDist(debate)
	chkr = SpellChecker("en_GB", debate)

	very_rare = []

	for w in longwords:
		if fdist[w] < 2 and chkr.check(w) == True:
			very_rare.append(w.encode('utf-8'))

	return very_rare
"""
def notso_rare_long_words(debate):             #takes lemmatised debate

	fdist = FreqDist(debate)

	notso_rare = []

	for w in debate:
		if fdist[w] > 3 and fdist[w] <6:
			notso_rare.append(w)

	return notso_rare

def not_rare_long_words(debate):             #takes lemmatised debate

	fdist = FreqDist(debate)

	not_rare = []

	for w in debate:	
		if fdist[w] > 5:
			not_rare.append(w)

	return not_rare
"""


###############################################################################################################################################



def lemmatised_words(debate):	
	
	lemmatised_words = debate
	#lemmatised_words = [lemma(w) for w in debate]
	#print lemmatised_words

	fdist = FreqDist(lemmatised_words)

	list_mc = fdist.most_common(20)

	most_common_lemmatised_list = []
	
	for list_tuple in list_mc:
		#print list_tuple[1]             #adjust for each!!
		most_common_lemmatised_list.append(list_tuple[0].encode('utf-8'))

	return most_common_lemmatised_list

#stemming using snowball stemmer
def stemmed_snowball(debate):

	snowball = SnowballStemmer("english")

	stemmed_words_snowball = [snowball.stem(w) for w in debate]

	fdist2 = FreqDist(stemmed_words_snowball)
	list_mc2 = fdist2.most_common(20)             #adjust this too!!!

	most_common_snowball = []
	
	for list_tuple in list_mc2:
		#print list_tuple[1]
		most_common_snowball.append(list_tuple[0].encode('utf-8'))

	return most_common_snowball



########################################################## most common nouns ##############################################################################

def extract_nouns(debate):
	#print debate
	"""
	debate = []
	for w in text:
		debate.append(w.lower())
	"""
	
	tagged = pos_tag(debate)
	#print tagged
	
	propernouns = [word for word,pos in tagged if pos == 'NNP']  #should be able to make or statement no?
	nnouns = [word for word,pos in tagged if pos == 'NN']
	nounss = [word for word,pos in tagged if pos == 'NNS']
	
	nouns = nnouns + propernouns + nounss

	c = Counter(nouns)
	most_common_nouns = c.most_common(5)            #adjustment for each debate!!!
	mcn_list = []
	for list_tuple in most_common_nouns:
		#print list_tuple[1]
		mcn_list.append(list_tuple[0])
	
	return mcn_list


#NOT NECESSARY - ENOUGH TO DO MOST COMMON WORDS!!!
"""
def extract_verbs(debate):
	tagged = pos_tag(debate)
	#print tagged
	
	v1 = [word for word,pos in tagged if pos == 'VBG']  #should be able to make or statement no?
	v2 = [word for word,pos in tagged if pos == 'VB']
	v3 = [word for word,pos in tagged if pos == 'VBD']
	v4 = [word for word,pos in tagged if pos == 'VBP']  #should be able to make or statement no?
	v5 = [word for word,pos in tagged if pos == 'VBN']
	v6 = [word for word,pos in tagged if pos == 'VBZ']
	
	verbs = v1 + v2 + v3 + v4 + v5 + v6

	c = Counter(verbs)
	most_common_verbs = c.most_common(30)
	mcv_list = []
	for list_tuple in most_common_verbs:
		word = list_tuple[0]
		print list_tuple[1]
		if len(word) > 4:
			mcv_list.append(list_tuple[0])
	
	return mcv_list



def extract_adjectives(debate):
	tagged = pos_tag(debate)
	#print tagged
	
	a1 = [word for word,pos in tagged if pos == 'JJ']  #should be able to make or statement no?
	a2 = [word for word,pos in tagged if pos == 'JJR']
	a3 = [word for word,pos in tagged if pos == 'JJS']
	
	adj = a1 + a2 + a3

	c = Counter(adj)
	most_common_adj = c.most_common(30)
	mca_list = []
	for list_tuple in most_common_adj:
		word = list_tuple[0]
		print list_tuple[1]
		if len(word) > 4:
			mca_list.append(list_tuple[0])
	
	return mca_list

"""

##########################################################    unusual stems   ##################################################################

def unusual_words(text, mcs, google):    
    text = text.split()

    #take text
    # spellcheck it
    # put correct words in list
    # stem them
    #check if word stem in one of the lists
    #return list of stems not in google list
    spellC = SpellChecker("en_GB")
    correct_words = []

    for w in text:
    	if w.isalpha():
    		correct = spellC.check(w)
    		if correct == True:
    			correct_words.append(w.encode('utf-8'))

    snowball = SnowballStemmer("english")

    delete_list = mcs + google

    stemmed_words = [snowball.stem(w) for w in correct_words]
    unusual_words = [w.encode('utf-8') for w in stemmed_words if w not in delete_list]
    unusual_words = set(unusual_words)
    unusual_words = list(unusual_words)
    
    return unusual_words                    #will miss words that are not in the enchant checker dictionary!!!!!! 



######################################################## named entity #########################################################

#if there is time i want to train a classifier, now its really bad

def preprocess_for_nee_and_print(argument):                        #need to write better function that might correct the sentence first before extracting them but for now it will do it
	

	sentences = nltk.sent_tokenize(argument)
	tokenized_sentences = [nltk.word_tokenize(sentence) for sentence in sentences]
	tagged_sentences = [nltk.pos_tag(sentence) for sentence in tokenized_sentences]
	chunked_sentences = nltk.ne_chunk_sents(tagged_sentences, binary=True)

	entity_names = []

	for tree in chunked_sentences:
		#tree.draw()
		entity_names.extend(extract_entity_names(tree))

	#print entity_names
	named_entities = [w for w in entity_names if w not in not_NE]
	#print entities
	#named_entities = [w for w in entities if w.lower() not in mcl]
	#named_entities = set(named_entities)
	#named_entities = list(named_entities)
	return named_entities

def extract_entity_names(tree):
    entity_names = []

    if hasattr(tree, 'label') and tree.label:
        if tree.label() == 'NE':
            entity_names.append(' '.join([child[0] for child in tree]))
        else:
            for child in tree:
                entity_names.extend(extract_entity_names(child))

    return entity_names
