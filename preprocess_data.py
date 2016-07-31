#PROCESS DATA

import openpyxl
import os
from whole_debate_func_clean import *
from ind_arguments import*


#**********************************************************************************************************************
"""right so this loops through the individual argument folder and analyses the debate and the individual arguments"""
#**********************************************************************************************************************

def demo():

	directory = '/home/user/Documents/IndProject/Data_and_Code/INDIVIDUAL_ARGUMENTS/'

	for file in os.listdir(directory):
		print "******************************************************************************************************************************"
		print "name of file:" , file
		whole_name = directory+file
		wb = openpyxl.load_workbook(whole_name)
		sheet = wb.get_sheet_by_name('Sheet1')
		end = sheet.max_row
		text = []

		#let the fun begin: first append all sentences to one big text
		for i in range(2,end,1):
			sentence = [sheet.cell(row = i, column =2).value.encode('utf-8')]
			text = text+sentence
		#print type(sentence)

		text= ' '.join(str(r) for r in text)
		#print text                              #great, works!

		###############################################################################################

		print "results whole debate: "
		tuple_hyp = preprocess_debateI(text)

		#number of hyperlinks mentioned in whole debate
		number_hyperlinks = tuple_hyp[1]
		print "hyperlinks: ", number_hyperlinks
		
		#no_punct is the whole debate without punctuation and tokenized in separate words
		no_punct = tuple_hyp[0]
		
		#all meaningful/relevant words
		no_stopwords = preprocess_debateII(no_punct)
		
		#most common lemmas and stems
		mc_lemma = lemmatised_words(no_stopwords)
		print "mc lemmas", mc_lemma

		mc_stem = stemmed_snowball(no_stopwords)
		print "mc stems", mc_stem

		#most common nouns
		nouns = extract_nouns(no_stopwords)
		print "mc nouns", nouns

		#named entities
		named_entities = preprocess_for_nee_and_print(text)
		print "ne", named_entities

		#unusual words compared to most common 10000 google words
		print "google"
		unusual_wordsI = unusual_words(text, mc_stem, google10_stemed)
		print unusual_wordsI

		#unusual words compared to most common 20000 google words
		unusual_wordsII = unusual_words(text, mc_stem, google20_stemed)
		print unusual_wordsII
		#######################################################################################################################
		"""
		print "_____________________________ individual arguments ____________________________________________________________"
		print "results sentence"
		for i in range(2, end, 1):
			#sentence aka argument in one cell
			sentence = sheet.cell(row = i, column =2).value.encode('utf-8')
			#################################################################################
			#super preprocessed for mc stuff extraction
			token_sentence = preprocess_argumentIV(sentence)         #sentence without stopwords
			
			#string without hyperlinks
			no_hyperlinks= preprocess_argumentI(sentence)
			#delete all punctuatuin apart from apostrophe (also string) 
			only_apostrophe_alllower = preprocess_argumentII(no_hyperlinks)

			#list
			all_lower_no_punct = preprocess_argumentIII(only_apostrophe_alllower)

			lemma_sent = [lemma(w) for w in token_sentence]
			snowball = SnowballStemmer("english")
			stem_sent  = [snowball.stem(w) for w in token_sentence]
			##################################################################################
			
			#need to delete stopwords and shit here too!!!!!!!! not just whole debate

			#linkingwordcount
			linkwords = linkingwordcount(all_lower_no_punct)                  #change formula - linking words compared to sentences not words in argument!
			print " linnking", linkwords

			readability = ColemanLiauIndex(no_hyperlinks)
			print "read", readability

			badword = badwordcount(all_lower_no_punct)                       #how do i count them?
			print "badwords",  badword

			spellcheck = spellCheck(only_apostrophe_alllower.split())
			print "errors" , spellcheck

			punctu = punctCount(sentence)
			print "punct" , punctu

			hyper = containsHyperlink(sentence)
			print "hyper", hyper

			caps = capsCount(sentence.split())                 #need all words
			print "caps", caps

			nee = preprocess_for_nee_and_print_ind(sentence)                 #need list for average, not set
			print "named entity", nee

			nouns_a = nouns_in_argument(token_sentence)                 #without stopwords!!                      
			print "nouns", nouns_a

			#intersection of nouns
			inter_noun = set(nouns).intersection(nouns_a)
			percent_inter = (float(len(inter_noun)) / len(nouns))
			print "percent intersection nouns", percent_inter

			#intersection of lemmas
			inter_lemmas = set(lemma_sent).intersection(mc_lemma)
			percent_interII = (float(len(inter_lemmas)) / len(mc_lemma))
			print "percent intersection lemmas", percent_interII

			#intersection of stems
			inter_stem = set(stem_sent).intersection(mc_stem)
			percent_interIII = (float(len(inter_stem)) / len(mc_stem))
			print "percent intersection stems", percent_interIII

			#intersection of unusual words!	
			"""




		######################################################################################################

		#could do the same for lemma and see if result is better but ill leave it like that for now!


	#print counter



if __name__ == '__main__':
    demo()
